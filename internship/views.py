"""
views.py
"""
from openpyxl import load_workbook
from django.shortcuts import render
from django.views.generic import TemplateView, ListView
from .models import Student, Internship_Assignment,Internship
from .forms import StudentSearchForm


class HomepageView(TemplateView):
    """
    template for home page
    """
    template_name = 'index.html'

class FileuploadView(TemplateView):
    """
    Template for uploading excel sheet to database
    """
    template_name = 'Upload.html'

    def import_data(request): # pylint: disable=too-many-locals
        """
        For importing data from excel sheet to database tables

        """
        book = load_workbook(filename = request, data_only=True)
        sheet = book.active
        student = []
        internship_assignment = []
        internship =[]
        student_id = 1
        intern_id  = 1
        intern_assing_id = 1
        for i in range(2,sheet.max_row+1):
            std = [student_id, sheet.cell(row=int(i), column=3).value,
                        sheet.cell(row=int(i), column=1).value,
                        sheet.cell(row=int(i), column=2).value,
                        sheet.cell(row=int(i), column=7).value,
                        sheet.cell(row=int(i), column=4).value,
                        sheet.cell(row=int(i), column=5).value,
                        sheet.cell(row=int(i), column=8).value]
            student_instance = Student(id, sheet.cell(row=int(i), column=3).value,
                        sheet.cell(row=int(i), column=1).value,
                        sheet.cell(row=int(i), column=2).value,
                        sheet.cell(row=int(i), column=7).value,
                        sheet.cell(row=int(i), column=4).value,
                        sheet.cell(row=int(i), column=5).value,
                        sheet.cell(row=int(i), column=8).value)
            student_instance.save()
            student.append(std)

            mailing_address = sheet.cell(row=int(i), column=20).value + " "
            city = sheet.cell(row=int(i), column=21).value + " "
            state = sheet.cell(row=int(i), column=22).value + " "
            zip_code = sheet.cell(row=int(i), column=23).value


            address = mailing_address + city + state + zip_code

            intern = [sheet.cell(row=int(i), column=14).value,
            sheet.cell(row=int(i), column=15).value,
            sheet.cell(row=int(i), column=18).value,
            sheet.cell(row=int(i), column=19).value,
            address,
            sheet.cell(row=int(i), column=24).value,
            sheet.cell(row=int(i), column=25).value,
            sheet.cell(row=int(i), column=26).value,
            sheet.cell(row=int(i), column=27).value]

            internship_instance = Internship(intern_id,sheet.cell(row=int(i), column=14).value,
            str(sheet.cell(row=int(i), column=15).value),
            sheet.cell(row=int(i), column=18).value,
            sheet.cell(row=int(i), column=19).value,
            address,
            sheet.cell(row=int(i), column=24).value,
            sheet.cell(row=int(i), column=25).value,
            sheet.cell(row=int(i), column=26).value,
            sheet.cell(row=int(i), column=27).value)
            internship_instance.save()
            internship.append(intern)

            intern_assign = [sheet.cell(row=int(i), column=9).value,
            str(sheet.cell(row=int(i), column=10).value),
            sheet.cell(row=int(i), column=11).value,
            str(sheet.cell(row=int(i), column=12).value),
            sheet.cell(row=int(i), column=13).value,
            str(sheet.cell(row=int(i), column=16).value),
            str(sheet.cell(row=int(i), column=17).value),]


            intern_assign = [sheet.cell(row=int(i), column=9).value,
            str(sheet.cell(row=int(i), column=10).value),
            sheet.cell(row=int(i), column=11).value,
            str(sheet.cell(row=int(i), column=12).value),
            sheet.cell(row=int(i), column=13).value,
            str(sheet.cell(row=int(i), column=16).value),
            str(sheet.cell(row=int(i), column=17).value),]

            internassign_instace = Internship_Assignment(intern_assing_id,student_id,intern_id,
            sheet.cell(row=int(i), column=9).value,
            str(sheet.cell(row=int(i), column=10).value),
            sheet.cell(row=int(i), column=11).value,
            str(sheet.cell(row=int(i), column=12).value),
            sheet.cell(row=int(i), column=13).value,
            str(sheet.cell(row=int(i), column=16).value),
            str(sheet.cell(row=int(i), column=17).value))

            internassign_instace.save()
            internship_assignment.append(intern_assign)
            student_id = student_id + 1
            intern_id = intern_id + 1
            intern_assing_id = intern_assing_id + 1

    def import_file(request):
        """
        getting file name from template
        """
        if request.method=='POST' and 'docfile' in request.FILES:
            files = request.FILES['docfile']
            FileuploadView.import_data(files)
        return render(request, 'Upload.html')

class StudentListView(ListView): # pylint: disable=too-many-ancestors
    """
    listing all the details of Student table
    """
    model = Student

    def display_students(request):
        """
        searching Student table based on first_name and last_name provided by user
        """
        button = "students"
        student_items = Student.objects.all()
        form = StudentSearchForm(request.POST or None)
        context = {
            'button' : button,
            'student_items' : student_items,
            'form' : form
        }
        if request.method == 'POST':
            student_items = Student.objects.filter(first_name__icontains=form['first_name'].value(),
                                              last_name__icontains=form['last_name'].value()
                                              )
            context = {
                "student_items" : student_items,
                "form": form
            }
        return render(request, 'index.html', context)



class Internship_AssignmentListView(ListView): # pylint: disable=too-many-ancestors
    """
    listing all the details of Internship_Assignment table
    """
    model = Internship_Assignment.objects.all()

class InternshipListView(ListView): # pylint: disable=too-many-ancestors
    """
    listing all the details of Internship table
    """
    model = Internship.objects.all()

def remove_all_data(request):
    """
    clearing the database tables
    """
    Student.objects.all().delete()
    Internship_Assignment.objects.all().delete()
    Internship.objects.all().delete()
    return render(request, 'base.html')
