from openpyxl import load_workbook
from ..models import Student, Internship_Assignment,Internship

def import_data(request): # pylint: disable=too-many-locals
    """
    For importing data from excel sheet to database tables
    """
    book = load_workbook(filename = request, data_only=True)
    sheet = book.active
    student = []
    internship_assignment = []
    internship =[]
    student_id = 1
    intern_id  = 1
    intern_assing_id = 1
    for i in range(2,sheet.max_row+1):
        std = [student_id, sheet.cell(row=int(i), column=3).value,
                    sheet.cell(row=int(i), column=1).value,
                    sheet.cell(row=int(i), column=2).value,
                    sheet.cell(row=int(i), column=7).value,
                    sheet.cell(row=int(i), column=4).value,
                    sheet.cell(row=int(i), column=5).value,
                    sheet.cell(row=int(i), column=8).value]
        student_instance = Student(student_id, sheet.cell(row=int(i), column=3).value,
                    sheet.cell(row=int(i), column=1).value,
                    sheet.cell(row=int(i), column=2).value,
                    sheet.cell(row=int(i), column=7).value,
                    sheet.cell(row=int(i), column=4).value,
                    sheet.cell(row=int(i), column=5).value,
                    sheet.cell(row=int(i), column=8).value)
        student_instance.save()
        student.append(std)

        mailing_address = sheet.cell(row=int(i), column=20).value + " "
        city = sheet.cell(row=int(i), column=21).value + " "
        state = sheet.cell(row=int(i), column=22).value + " "
        zip_code = str(sheet.cell(row=int(i), column=23).value)


        address = mailing_address + city + state + zip_code

        intern = [sheet.cell(row=int(i), column=14).value,
        sheet.cell(row=int(i), column=15).value,
        sheet.cell(row=int(i), column=18).value,
        sheet.cell(row=int(i), column=19).value,
        address,
        sheet.cell(row=int(i), column=24).value,
        sheet.cell(row=int(i), column=25).value,
        sheet.cell(row=int(i), column=26).value,
        sheet.cell(row=int(i), column=27).value]

        internship_instance = Internship(intern_id,sheet.cell(row=int(i), column=14).value,
        str(sheet.cell(row=int(i), column=15).value),
        sheet.cell(row=int(i), column=18).value,
        sheet.cell(row=int(i), column=19).value,
        address,
        sheet.cell(row=int(i), column=24).value,
        sheet.cell(row=int(i), column=25).value,
        sheet.cell(row=int(i), column=26).value,
        sheet.cell(row=int(i), column=27).value)
        internship_instance.save()
        internship.append(intern)

        intern_assign = [sheet.cell(row=int(i), column=9).value,
        str(sheet.cell(row=int(i), column=10).value),
        sheet.cell(row=int(i), column=11).value,
        str(sheet.cell(row=int(i), column=12).value),
        sheet.cell(row=int(i), column=13).value,
        str(sheet.cell(row=int(i), column=16).value),
        str(sheet.cell(row=int(i), column=17).value),]


        intern_assign = [sheet.cell(row=int(i), column=9).value,
        str(sheet.cell(row=int(i), column=10).value),
        sheet.cell(row=int(i), column=11).value,
        str(sheet.cell(row=int(i), column=12).value),
        sheet.cell(row=int(i), column=13).value,
        str(sheet.cell(row=int(i), column=16).value),
        str(sheet.cell(row=int(i), column=17).value),]

        internassign_instace = Internship_Assignment(intern_assing_id,student_id,intern_id,
        sheet.cell(row=int(i), column=9).value,
        str(sheet.cell(row=int(i), column=10).value),
        sheet.cell(row=int(i), column=11).value,
        str(sheet.cell(row=int(i), column=12).value),
        sheet.cell(row=int(i), column=13).value,
        str(sheet.cell(row=int(i), column=16).value),
        str(sheet.cell(row=int(i), column=17).value))

        internassign_instace.save()
        internship_assignment.append(intern_assign)
        student_id = student_id + 1
        intern_id = intern_id + 1
        intern_assing_id = intern_assing_id + 1
